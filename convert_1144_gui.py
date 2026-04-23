"""
GUI-обёртка для convert_1144.py
Упаковка в exe: pip install pyinstaller && pyinstaller --onefile --windowed convert_1144_gui.py
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import os
import sys
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import re

# ── Вся логика конвертации (из convert_1144.py) ─────────────────────────────

VAT = 1.22

SHEET_CONFIG = {
    'Полки':                              ('Яндекс Маркет', 'Показы'),
    'Буст продаж, оплата за показы':      ('Яндекс Маркет', 'Показы'),
    'Буст продаж, оплата за продажи':     ('Яндекс Маркет', 'Лиды'),
    'Поп-ап уведомления (аукцион)':       ('Яндекс Маркет', 'Лиды'),
    'Рекламно-информационные материа':    ('Яндекс Маркет', 'Лиды'),
    'Реклама на внешних площадках':       ('Яндекс Маркет', 'Лиды'),
    'Баннеры (аукцион)':                  ('Яндекс Лавка',  'Пакеты'),
    'Баннеры кликаут (аукцион)':          ('Яндекс Лавка',  'Пакеты'),
    'Полки с оплатой по дням':            ('Яндекс Лавка',  'Пакеты'),
    'Реклама на внешних площадках-М':     ('Яндекс Лавка',  'Пакеты'),
    'Организация и проведение маркет':    ('Яндекс Маркет', 'Пакеты'),
}

SKIP_SHEETS = {
    'Баннеры (бронирование)', 'Баннеры кликаут (бронирование)',
    'Пуши (бронирование)', 'Пуши (аукцион)',
    'Поп-ап уведомления (бронировани', 'Реклама в планшетах (бронирован)',
    'Реклама в планшетах (аукцион)', 'ТВ реклама',
    'Информационная рассылка по зада', 'Специальные размещения', 'Отзывы за баллы',
}

JDE_CABINETS = {"mdp_jacobs", "mdp_carte noire", "mdp_l'or"}

def is_jde_cabinet(cabinet_name):
    if pd.isna(cabinet_name):
        return True
    return str(cabinet_name).strip().lower() in JDE_CABINETS

NON_JDE_KEYWORDS = ['Sandoz', 'Linex', 'Red Bull', 'Dobry', 'ИФЛ', 'Ингосстрах']

def is_jde_campaign(campaign_name):
    name = str(campaign_name)
    return not any(kw.lower() in name.lower() for kw in NON_JDE_KEYWORDS)

def detect_brand(campaign_name):
    name = str(campaign_name)
    priority = [
        ('NCC CARTE NOIRE',      'NCC CARTE NOIRE'),
        ('CARTE NOIRE INSTANT',  'CARTE NOIRE INSTANT'),
        ('Carte Noire Instant',  'CARTE NOIRE INSTANT'),
        ('Carte Noire  Instant', 'CARTE NOIRE INSTANT'),
        ('CARTE NOIRE MUSE',     'CARTE NOIRE MUSE'),
        ('Carte Noire MuSe',     'CARTE NOIRE MUSE'),
        ('Carte Noire_MuSe',     'CARTE NOIRE MUSE'),
        ('CARTE NOIRE MIXES',    'CARTE NOIRE MIXES'),
        ('CARTE NOIRE',          'CARTE NOIRE INSTANT'),
        ('NCC MONARCH',          'NCC MONARCH'),
        ('Monarch_NCC',          'NCC MONARCH'),
        ('Monarch NCC',          'NCC MONARCH'),
        ('Monarch_Ncc',          'NCC MONARCH'),
        ('MONARCH_NCC',          'NCC MONARCH'),
        ('MONARCH INSTANT',      'MONARCH INSTANT'),
        ('Monarch Instant',      'MONARCH INSTANT'),
        ('MONARCH MUSE',         'MONARCH MUSE'),
        ('Monarch MuSe',         'MONARCH MUSE'),
        ('Monarch MUSE',         'MONARCH MUSE'),
        ('MONARCH MIXES',        'MONARCH MIXES'),
        ('Monarch Mixes',        'MONARCH MIXES'),
        ("NCC L\u2019OR",        "NCC L'OR"),
        ("NCC L'OR",             "NCC L'OR"),
        ("L'OR NCC",             "NCC L'OR"),
        ("L\u2019OR NCC",        "NCC L'OR"),
        ("L`OR",                 "NCC L'OR"),
        ("L'OR",                 "NCC L'OR"),
        ('NCC L',                "NCC L'OR"),
        ('дед/слоу',             'MONARCH INSTANT'),
        ('Monarch',              'MONARCH INSTANT'),
    ]
    for key, brand in priority:
        if key.lower() in name.lower():
            return brand
    return None

def detect_brand_from_parts(campaign_raw):
    parts = [p.strip() for p in re.split(r'//|\\\\+', str(campaign_raw))]
    for part in parts[1:]:
        brand = detect_brand(part)
        if brand:
            return brand
    return detect_brand(campaign_raw)

def detect_campaign_name(campaign_raw, sheet_name):
    name = str(campaign_raw).lower()
    raw = str(campaign_raw)

    if sheet_name == 'Полки':
        if 'карточка товара' in name: return 'Брендовая полка_карточка товара'
        if 'карточка' in name:        return 'Брендовая полка_карточка'
        if 'поиск' in name:           return 'Брендовая полка_поиск'
        if 'зерно' in name or 'баннер' in name: return 'Брендовая полка_зерно_баннер'
        if 'молотый' in name:         return 'Брендовая полка_молотый'
        return 'Брендовая полка'
    if sheet_name == 'Буст продаж, оплата за показы':
        return 'Буст показов'
    if sheet_name == 'Буст продаж, оплата за продажи':
        return 'буст продаж'
    if sheet_name == 'Поп-ап уведомления (аукцион)':
        m = re.search(r'\d{2}\.\d{2}\.\d{4}', raw)
        return f'Поп-ап {m.group(0)}' if m else 'Поп-ап'
    if sheet_name in ('Рекламно-информационные материа', 'Реклама на внешних площадках'):
        parts = raw.split('_')
        clean = [p for p in parts if p not in ('FMCG', 'JDE')]
        if clean and detect_brand(clean[0]):
            clean = clean[1:]
        return '_'.join(clean) if clean else raw
    if sheet_name in ('Баннеры (аукцион)', 'Баннеры кликаут (аукцион)'):
        parts = [p.strip() for p in re.split(r'\\\\+', raw)]
        if len(parts) >= 3:
            return parts[2].strip()
        return raw
    if sheet_name == 'Полки с оплатой по дням':
        if 'бренд-полка' in name or 'бренд полка' in name:
            return 'Отдельная брендовая полка в категории'
        if 'товар - герой' in name or 'товар-герой' in name:
            return 'Товар-герой в поиске'
        parts = [p.strip() for p in raw.split('//')]
        return parts[2] if len(parts) >= 3 else raw
    if sheet_name == 'Реклама на внешних площадках-М':
        parts = [p.strip() for p in re.split(r'//|\\\\', raw)]
        camp = re.sub(r'^\d+\.\s*', '', parts[0].strip())
        return camp
    if sheet_name == 'Организация и проведение маркет':
        parts = raw.split('_')
        while parts and parts[0] in ('FMCG', 'JDE'):
            parts = parts[1:]
        brand_kw = ['carte noire', 'monarch', 'ncc', 'jacobs']
        if parts and any(kw in parts[0].lower() for kw in brand_kw) \
                and 'промокод' not in parts[0].lower() \
                and 'слоу' not in parts[0].lower() \
                and 'дед' not in parts[0].lower():
            parts = parts[1:]
        return '_'.join(parts) if parts else raw
    return raw


def detect_mechanic_from_type(service_type, default_mechanic):
    """Определяет механику по колонке Тип услуги."""
    if pd.isna(service_type):
        return default_mechanic
    st = str(service_type).strip()
    if 'Лавка' in st:
        return 'Яндекс Лавка'
    if 'Еда' in st:
        return 'Яндекс Еда'
    return default_mechanic

def read_sheet_data(df):
    for i, row in df.iterrows():
        if str(row.iloc[0]).strip() == 'Дата':
            data = df.iloc[i+1:].copy()
            data.columns = df.iloc[i].tolist()
            data = data[data['Дата'].notna() & (data['Дата'].astype(str).str.strip() != 'Итого:')]
            data['Дата'] = pd.to_datetime(data['Дата'], errors='coerce')
            data['Сумма (с НДС), ₽'] = pd.to_numeric(data['Сумма (с НДС), ₽'], errors='coerce')
            return data.dropna(subset=['Дата'])
    return pd.DataFrame()

def process_file(input_path, output_path, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)

    log(f"Читаем: {os.path.basename(input_path)}")
    sheets = pd.read_excel(input_path, sheet_name=None, header=None)
    log(f"Найдено листов: {len(sheets)}")

    all_data = []
    warnings = []

    for sheet_name, raw_df in sheets.items():
        if sheet_name in SKIP_SHEETS:
            continue
        if sheet_name not in SHEET_CONFIG:
            warnings.append(f"Неизвестный лист '{sheet_name}' — пропущен")
            continue

        mechanic, default_buy_type = SHEET_CONFIG[sheet_name]
        df = read_sheet_data(raw_df)
        if df.empty or 'Кампания' not in df.columns:
            continue

        df = df[df['Кампания'].apply(is_jde_campaign)]
        if 'Название кабинета' in df.columns:
            df = df[df['Название кабинета'].apply(is_jde_cabinet)]
        if df.empty:
            continue

        grouped = df.groupby('Кампания').agg(
            total=('Сумма (с НДС), ₽', 'sum'),
            date_min=('Дата', 'min'),
            date_max=('Дата', 'max'),
        ).reset_index()

        # Для листов с Тип услуги — группируем с учётом механики
        has_service_type = 'Тип услуги' in df.columns

        for _, row in grouped.iterrows():
            camp_raw = row['Кампания']
            budget = row['total'] / VAT
            period_start = row['date_min'].to_pydatetime() if pd.notna(row['date_min']) else datetime.now()
            period_end   = row['date_max'].to_pydatetime() if pd.notna(row['date_max']) else datetime.now()

            # Определяем механику по Тип услуги если есть
            if has_service_type and sheet_name == 'Реклама на внешних площадках-М':
                service_types = df[df['Кампания'] == camp_raw]['Тип услуги']
                service_type = service_types.dropna().iloc[0] if not service_types.dropna().empty else None
                actual_mechanic = detect_mechanic_from_type(service_type, mechanic)
            else:
                actual_mechanic = mechanic

            if sheet_name in ('Реклама на внешних площадках-М', 'Баннеры (аукцион)',
                              'Баннеры кликаут (аукцион)', 'Полки с оплатой по дням'):
                brand = detect_brand_from_parts(camp_raw)
            else:
                brand = detect_brand(camp_raw)

            if not brand:
                warnings.append(f"Не определён бренд: '{camp_raw}'")
                brand = camp_raw

            camp_name = detect_campaign_name(camp_raw, sheet_name)
            all_data.append({
                'brand': brand, 'campaign': camp_name,
                'mechanic': actual_mechanic, 'buy_type': default_buy_type,
                'budget': budget, 'period_start': period_start, 'period_end': period_end,
            })

    # Строим Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Лист1'

    headers = [
        'Клиент', 'Бренд', 'Кампания (название)', 'Отдел', 'Направление',
        'Начало периода', 'Конец периода', 'Инвентарь', 'Механика\n(площадка)',
        'Тип закупки', 'Количество\nПлан', 'Стоимость\nПлан', 'Плановый бюджет',
        'Закрытие план/факт ', 'Фактический объем ', 'Фактический бюджет',
        'К оплате поставщику', 'Фактический расход', 'Комментарий '
    ]

    hf = Font(name='Arial', bold=True, size=10)
    hfill = PatternFill('solid', start_color='D9E1F2')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cf = Font(name='Arial', size=10)
    ca = Alignment(horizontal='left', vertical='center')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ha; cell.border = border
    ws.row_dimensions[1].height = 30

    for row_idx, item in enumerate(all_data, 2):
        values = [
            'JDE', item['brand'], item['campaign'], 'Performance', 'Marketplaces',
            item['period_start'], item['period_end'], 'Прямая закупка',
            item['mechanic'], item['buy_type'],
            1, 1000, None, 'Факт', None,
            str(item['budget']).replace(',', '.'),
            str(item['budget']).replace(',', '.'),
            None, 'Передача данных В ЕРИР - нет',
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = cf; cell.alignment = ca; cell.border = border
            if col in (6, 7) and val is not None:
                cell.number_format = 'dd\\.mm\\.yyyy'

    col_widths = [8, 22, 45, 14, 14, 16, 16, 16, 16, 12, 14, 14, 16, 18, 18, 18, 20, 18, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    wb.save(output_path)

    log(f"Строк обработано: {len(all_data)}")
    for w in warnings:
        log(f"  ⚠ {w}")
    log(f"✓ Готово: {os.path.basename(output_path)}")

# ── GUI ───────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("JDE 1144 Конвертер")
        self.resizable(False, False)
        self.configure(bg="#1a1a2e")
        self._build_ui()
        self._center()

    def _center(self):
        self.update_idletasks()
        w, h = 540, 400
        x = (self.winfo_screenwidth() - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _build_ui(self):
        PAD = 24
        BG  = "#1a1a2e"
        CARD = "#16213e"
        ACC = "#e94560"
        FG  = "#eaeaea"
        MUTED = "#8892a4"

        # Заголовок
        tk.Label(self, text="JDE 1144", font=("Georgia", 22, "bold"),
                 bg=BG, fg=ACC).pack(pady=(PAD, 2))
        tk.Label(self, text="Конвертер рекламных отчётов Яндекс Маркет",
                 font=("Georgia", 10), bg=BG, fg=MUTED).pack(pady=(0, PAD))

        # Карточка выбора файла
        card = tk.Frame(self, bg=CARD, padx=20, pady=16)
        card.pack(fill="x", padx=PAD)

        tk.Label(card, text="Исходный файл (.xlsx)", font=("Consolas", 9),
                 bg=CARD, fg=MUTED).pack(anchor="w")

        row = tk.Frame(card, bg=CARD)
        row.pack(fill="x", pady=(4, 0))

        self.path_var = tk.StringVar(value="Файл не выбран")
        tk.Label(row, textvariable=self.path_var, font=("Consolas", 9),
                 bg=CARD, fg=FG, anchor="w", width=42,
                 relief="flat", bd=0).pack(side="left", fill="x", expand=True)

        tk.Button(row, text="Обзор…", font=("Consolas", 9), bg=ACC, fg="white",
                  activebackground="#c73652", activeforeground="white",
                  relief="flat", padx=10, cursor="hand2",
                  command=self._browse).pack(side="right")

        # Кнопка запуска
        self.btn = tk.Button(self, text="▶  Конвертировать",
                             font=("Georgia", 12, "bold"),
                             bg=ACC, fg="white",
                             activebackground="#c73652", activeforeground="white",
                             relief="flat", padx=20, pady=8,
                             cursor="hand2", command=self._run)
        self.btn.pack(pady=PAD)

        # Прогресс
        self.progress = ttk.Progressbar(self, mode="indeterminate", length=490)
        style = ttk.Style(self)
        style.theme_use("default")
        style.configure("TProgressbar", troughcolor=CARD, background=ACC, thickness=4)
        self.progress.pack(padx=PAD)

        # Лог
        log_frame = tk.Frame(self, bg=CARD)
        log_frame.pack(fill="both", expand=True, padx=PAD, pady=(8, PAD))

        self.log_box = tk.Text(log_frame, height=7, font=("Consolas", 9),
                               bg=CARD, fg=FG, relief="flat",
                               state="disabled", wrap="word",
                               insertbackground=FG)
        self.log_box.pack(fill="both", expand=True, padx=8, pady=8)

        self._input_path = None

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Выберите файл",
            filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")]
        )
        if path:
            self._input_path = path
            self.path_var.set(os.path.basename(path))

    def _log(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _run(self):
        if not self._input_path:
            messagebox.showwarning("Файл не выбран", "Пожалуйста, выберите исходный файл.")
            return

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        base = os.path.splitext(os.path.basename(self._input_path))[0]
        output_path = os.path.join(desktop, f"{base}_результат.xlsx")

        self.btn.configure(state="disabled")
        self.progress.start(10)
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

        def task():
            try:
                process_file(self._input_path, output_path, log_callback=self._log)
                self.after(0, lambda: messagebox.showinfo(
                    "Готово",
                    f"Файл сохранён на рабочий стол:\n{os.path.basename(output_path)}"
                ))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Ошибка", str(e)))
                self._log(f"ОШИБКА: {e}")
            finally:
                self.after(0, self.progress.stop)
                self.after(0, lambda: self.btn.configure(state="normal"))

        threading.Thread(target=task, daemon=True).start()


if __name__ == "__main__":
    app = App()
    app.mainloop()
